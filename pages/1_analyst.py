import streamlit as st
import os
import polars as pl
import plotly.express as px
import pandas as pd
import sys
from pathlib import Path

from run_fast_pipeline import run_fast_processing
from run_pipeline import run_full_slow_pipeline

PROJECT_ROOT = Path(__file__).resolve().parent.parent   
sys.path.insert(0, str(PROJECT_ROOT))
from src.core.analytics import IncidentAnalytics

st.set_page_config(page_title="Analyst | Анализатор инцидентов", layout="wide")
st.title("📊 Система анализа инцидентов ЖКХ и муниципалитетов")

TEMP_RAW_PATH = "data/raw/temp_uploaded.xlsx"
PROCESSED_XLSX_PATH = "data/processed/тестовый_файл_slow.xlsx"
OUTPUT_TXT_REPORT = "data/output/report.txt"
OUTPUT_XLSX_REPORT = "data/output/top_municipalities.xlsx"

st.sidebar.header("Параметры обработки")

test_file_fast = "data/raw/тестовый файл.xlsx"
test_file_slow = "data/raw/test_40.xlsx"
has_test_fast = os.path.exists(test_file_fast)
has_test_slow = os.path.exists(test_file_slow)

use_test_file = False
if has_test_fast or has_test_slow:
    st.sidebar.success("✅ Тестовые файлы найдены")
    use_test_file = st.sidebar.checkbox("Использовать тестовые файлы", value=False)

uploaded_file = None
if not use_test_file:
    uploaded_file = st.sidebar.file_uploader("Загрузите исходный файл (.xlsx)", type=["xlsx"], key="analyst_upload")

pipeline_mode = st.sidebar.radio(
    "Выберите режим обработки:", 
    ["Базовое решение (fast)", "AI-решение (slow)"]
)

execute_button = st.sidebar.button("Запустить анализ", key="analyst_run")

if use_test_file or uploaded_file is not None:
    os.makedirs(os.path.dirname(TEMP_RAW_PATH), exist_ok=True)
    
    if use_test_file:
        if pipeline_mode == "Базовое решение (fast)" and has_test_fast:
            import shutil
            shutil.copy(test_file_fast, TEMP_RAW_PATH)
            st.sidebar.success(f"✅ Используется тестовый файл.xlsx")
        elif pipeline_mode == "AI-решение (slow)" and has_test_slow:
            from openpyxl import load_workbook
            import shutil
            
            wb = load_workbook(test_file_slow)
            ws = wb.active
            ws.delete_rows(11, ws.max_row) 
            os.makedirs(os.path.dirname(TEMP_RAW_PATH), exist_ok=True)
            wb.save(TEMP_RAW_PATH)
            st.sidebar.success(f"✅ Используется small_data.xlsx")
    else:
        with open(TEMP_RAW_PATH, "wb") as f:
            f.write(uploaded_file.getbuffer())
        
    if execute_button:
        with st.spinner("Выполняется предобработка данных, расчет метрик и генерация отчетов..."):
            if pipeline_mode == "Базовое решение (fast)":
                run_fast_processing(input_path=TEMP_RAW_PATH, output_path=PROCESSED_XLSX_PATH)
            else:
                run_full_slow_pipeline(TEMP_RAW_PATH, PROCESSED_XLSX_PATH)
                
            df_processed = pl.read_excel(PROCESSED_XLSX_PATH, engine="calamine")
            analytics = IncidentAnalytics(df_processed)
            
            top_regions = analytics.build_reports(
                txt_output_path=OUTPUT_TXT_REPORT,
                xlsx_output_path=OUTPUT_XLSX_REPORT
            )
            
            st.session_state["top_municipalities"] = top_regions
            st.session_state["analyzer"] = analytics
            st.session_state["processing_done"] = True
            
            st.success("Анализ успешно завершен!")

    if st.session_state.get("processing_done"):
        st.subheader("Скачать результаты анализа")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            with open(OUTPUT_TXT_REPORT, "r", encoding="utf-8") as f:
                st.download_button(
                    label="Скачать текстовый отчет (Ollama AI) .txt",
                    data=f.read(),
                    file_name="Аналитический_отчет.txt",
                    mime="text/plain"
                )
        with col2:
            with open(OUTPUT_XLSX_REPORT, "rb") as f:
                st.download_button(
                    label="Скачать таблицы Топ-3 / Топ-10 .xlsx",
                    data=f.read(),
                    file_name="Рейтинг_муниципалитетов.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        with col3:
            if os.path.exists(PROCESSED_XLSX_PATH):
                with open(PROCESSED_XLSX_PATH, "rb") as f:
                    st.download_button(
                        label="Скачать обработанные данные .xlsx",
                        data=f.read(),
                        file_name="Обработанные_данные.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
        
        st.divider()
        
        st.subheader("Визуализация структуры проблем по муниципалитетам")
        
        selected_mun = st.selectbox(
            "Выберите муниципалитет для детализации:",
            st.session_state["top_municipalities"],
            key="analyst_mun_select"
        )
        
        if selected_mun:
            chart_df = st.session_state["analyzer"].get_region_distribution(selected_mun)

            if not chart_df.is_empty():
                pandas_chart_data = chart_df.to_pandas()

                name_col = "Teма" if "Teма" in pandas_chart_data.columns else "Тема"
                pandas_chart_data = pandas_chart_data.rename(columns={name_col: "topic"})

                pandas_chart_data["count"] = pd.to_numeric(pandas_chart_data["count"], errors="coerce").fillna(0)

                total = pandas_chart_data["count"].sum()
                if total > 0:
                    pandas_chart_data["pct"] = pandas_chart_data["count"] / total

                    pandas_chart_data.loc[pandas_chart_data["pct"] < 0.02, "topic"] = "Другое"

                    grouped = (
                        pandas_chart_data
                        .groupby("topic", as_index=False)["count"]
                        .sum()
                    )

                    fig = px.pie(
                        grouped,
                        values="count",
                        names="topic",
                        title=f"Соотношение типов проблем в: {selected_mun}",
                        hole=0.3,
                    )
                    fig.update_traces(
                        textinfo="percent",
                        textposition="inside",
                        showlegend=True
                    )
                    fig.update_layout(margin=dict(l=20, r=20, t=40, b=20))
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.info("В выбранном регионе суммарное количество инцидентов равно нулю.")
            else:
                st.info("В выбранном регионе отсутствуют зарегистрированные инциденты.")

        st.subheader("📊 Топ-10 муниципалитетов по суммарной критичности")
        
        top_regions_df = st.session_state["analyzer"].get_top_regions(limit=10)
        if not top_regions_df.is_empty():
            top_regions_pandas = top_regions_df.to_pandas()
            
            fig_top = px.bar(
                top_regions_pandas.sort_values("Суммы баллов", ascending=True),
                x="Суммы баллов",
                y="Муниципалитет",
                orientation="h",
                title="Рейтинг муниципалитетов по критичности проблем",
                labels={"Суммы баллов": "Критичность (баллы)", "Муниципалитет": ""},
                color="Суммы баллов",
                color_continuous_scale="Viridis"
            )
            fig_top.update_layout(
                height=400,
                margin=dict(l=150, r=20, t=40, b=20),
                coloraxis_colorbar=dict(title="Баллы")
            )
            st.plotly_chart(fig_top, use_container_width=True)
        
        st.divider()
else:
    st.info("Пожалуйста, загрузите входной файл формата .xlsx в боковую панель для начала работы или используйте тестовые файлы.")
