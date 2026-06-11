import streamlit as st
import os
import pandas as pd
import re
import sys
from pathlib import Path
from openpyxl import load_workbook

st.markdown("""
<style>
            
.page-header{
    margin-bottom:2rem;
}

.page-badge{
    display:inline-block;
    padding:6px 12px;
    border-radius:999px;
    background:rgba(15,98,254,.08);
    color:#0f62fe;
    font-size:.75rem;
    font-weight:700;
    letter-spacing:.08em;
    margin-bottom:.8rem;
}

.title-row{
    display:flex;
    align-items:center;
    gap:14px;
}

.title-bar{
    width:5px;
    height:44px;
    background:#0f62fe;
    border-radius:3px;
    flex-shrink:0;
}

.page-title{
    font-size:2.2rem;
    font-weight:700;
    margin:0;
    line-height:1.1;
}

.page-subtitle{
    margin-top:.8rem;
    color:gray;
    font-size:1rem;
    padding-bottom:1rem;
    border-bottom:1px solid rgba(120,120,120,.2);
}

/* =========================
   THEME VARIABLES (LIGHT + DARK SAFE)
========================= */
:root {
    --wf-bg: var(--secondary-background-color);
    --wf-border: rgba(15,98,254,.12);
    --wf-text: var(--text-color);
    --wf-muted: rgba(120,120,120,.85);
    --wf-accent: #0f62fe;
}

/* =========================
   CONTAINER
========================= */
.workflow-container{
    display:flex;
    gap:1rem;
    flex-wrap:wrap;
    margin: 1rem 0 2rem 0;
}

/* =========================
   STEP CARD (MAIN ELEMENT)
========================= */
.workflow-step-hz{
    flex:1;
    min-width:180px;

    background:var(--wf-bg);
    border:1px solid var(--wf-border);
    border-radius:14px;

    padding:1.1rem 1rem;

    transition: all .25s ease;
    position:relative;

    box-shadow: 0 2px 8px rgba(0,0,0,.04);
}

/* hover effect */
.workflow-step-hz:hover{
    transform: translateY(-3px);
    border-color: rgba(15,98,254,.35);
    box-shadow: 0 10px 24px rgba(15,98,254,.10);
}

/* =========================
   NUMBER BADGE
========================= */
.step-num-hz{
    display:inline-block;

    font-size:0.72rem;
    font-weight:700;
    letter-spacing:0.08em;
    text-transform:uppercase;

    color:var(--wf-accent);

    margin-bottom:0.6rem;
}

/* =========================
   TITLE
========================= */
.step-title-hz{
    font-size:0.98rem;
    font-weight:700;

    margin-bottom:0.3rem;

    color:var(--wf-text);
}

/* =========================
   DESCRIPTION
========================= */
.step-desc-hz{
    font-size:0.85rem;
    line-height:1.4;

    color:var(--wf-muted);
}

/* =========================
   OPTIONAL: CONNECTOR LINE FEEL (VISUAL FLOW)
========================= */
.workflow-step-hz:not(:last-child)::after{
    content:"";
    position:absolute;
    right:-0.5rem;
    top:50%;
    width:1rem;
    height:2px;
    background:rgba(15,98,254,.2);
    display:none; /* включи, если хочешь "цепочку" */
}
            
.step-card{
    flex:1;
    min-width:180px;

    background:var(--wf-bg);
    border:1px solid var(--wf-border);
    border-radius:14px;

    padding:1.1rem 1rem;

    transition: all .25s ease;
    position:relative;

    box-shadow: 0 2px 8px rgba(0,0,0,.04);
}

.step-number{
    display:inline-block;
    font-size:0.72rem;
    font-weight:700;
    color:#0f62fe;
    margin-bottom:0.5rem;
}

</style>
""", unsafe_allow_html=True)

PROJECT_ROOT = Path(__file__).resolve().parent.parent   
sys.path.insert(0, str(PROJECT_ROOT))

from src.llm.llama_client import generate

from src.workspace import render_workspace_panel, init_workspace, get_workspace, set_workspace_file, FileManager

st.set_page_config(page_title="Generator | Анализатор инцидентов", layout="wide")

st.markdown("""
<div class="page-header">

<div class="page-badge">
RESPONSE GENERATOR MODULE
</div>

<div class="title-row">
    <div class="title-bar"></div>
    <h1 class="page-title">
        Генератор ответов на обращения
    </h1>
</div>

<div class="page-subtitle">
LLM формирует официальные черновики ответов на обращения граждан на основе классификации инцидентов
</div>

</div>
""", unsafe_allow_html=True)

init_workspace()

left_col, right_col = st.columns([3, 1])

with right_col:
    render_workspace_panel(active_page="generator")

with left_col:
    fm = FileManager()

    TEMPLATE_EXAMPLES = """
    Пример 1:
    Жалоба: Здравствуйте, на улице Пушкина уже неделю не чистят снег, тротуары завалены, пройти невозможно.
    Ответ: Здравствуйте. Ваше обращение по вопросу уборки снега на ул. Пушкина получено. Информация передана в дорожную службу для проведения работ по очистке. Приносим извинения за доставленные неудобства.

    Пример 2:
    Жалоба: В квартире холодно, батареи еле теплые, температура в комнате +14 градусов.
    Ответ: Здравствуйте. Ваше обращение по вопросу низкой температуры в квартире зарегистрировано. Для проведения проверки системы отопления просим уточнить адрес и контактные данные. Обращение будет рассмотрено в установленном порядке.

    Пример 3:
    Жалоба: Добрый день, мусор не вывозят уже вторую неделю, контейнеры переполнены.
    Ответ: Здравствуйте. Ваше обращение по вопросу вывоза мусора получено. Информация передана региональному оператору для корректировки графика. Принимаются меры в рамках компетенции.
    """

    def load_excel_simple(file_path: str) -> pd.DataFrame:
        """Загружает Excel файл с любой структурой столбцов"""
        wb = load_workbook(file_path, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(x).strip() if x else "" for x in rows[0]]
        data = []
        for row in rows[1:]:
            record = {h: row[i] if i < len(row) else None for i, h in enumerate(headers)}
            data.append(record)
        return pd.DataFrame(data)

    def clean_text(text):
        """Очищает текст от специальных символов и форматирования"""
        if pd.isna(text) or not isinstance(text, str):
            return ""
        text = re.sub(r'\[id\d+\|([^\]]+)\]', r'\1', text)
        text = re.sub(r'\[club\d+\|([^\]]+)\]', r'\1', text)
        text = re.sub(r'@\w+', '', text)
        text = re.sub(r'https?://\S+', '', text)
        text = re.sub(r'<[^>]+>', '', text)
        text = re.sub(r'\s+', ' ', text)
        return text.strip()

    def remove_non_russian(text):
        """Удаляет не-русские символы (иероглифы и т.д.)"""
        if not isinstance(text, str):
            return text
        text = re.sub(r'[\u4e00-\u9fff\u3400-\u4dbf\uf900-\ufaff]', '', text)
        text = re.sub(r'[\u3040-\u309f\u30a0-\u30ff]', '', text)
        text = re.sub(r'[\uac00-\ud7af]', '', text)
        text = re.sub(r'\s+', ' ', text)
        return text.strip()

    def build_prompt(complaint_text, topic_group, subtopic, department, municipality):
        """Формирует промпт для генерации персонализированного черновика"""
        prompt = f"""<s>Ты — сотрудник администрации города Омска. Отвечай ТОЛЬКО на русском языке. Никаких других языков.

    Напиши черновик ответа на жалобу жителя.

    Правила:
    1. ОТВЕЧАЙ ТОЛЬКО НА РУССКОМ ЯЗЫКЕ
    2. Упомяни суть проблемы из жалобы (адрес, что именно случилось) — это персонализация
    3. НЕ придумывай конкретные даты, сроки, имена сотрудников, номера документов
    4. НЕ обещай конкретных результатов — только "принято в работу", "передано для рассмотрения"
    5. Ответ должен быть вежливым и официальным
    6. Это черновик — реальный сотрудник добавит конкретику

    Примеры:

    {TEMPLATE_EXAMPLES}

    Теперь напиши черновик ответа на эту жалобу. ТОЛЬКО НА РУССКОМ ЯЗЫКЕ.

    Тема: {topic_group}
    Категория: {subtopic}
    Отдел: {department}
    Район: {municipality}

    Текст жалобы: {complaint_text[:500]}

    Черновик ответа (на русском):"""
        return prompt

    st.markdown("""
    <div style="
        padding: 1.2rem 1.4rem;
        border: 1px solid rgba(15,98,254,.25);
        border-left: 5px solid #0f62fe;
        border-radius: 10px;
        background: var(--secondary-background-color);
        margin: 1rem 0 1.5rem 0;
    ">

    <div style="
        font-size: 1.1rem;
        font-weight: 700;
        margin-bottom: 0.6rem;
    ">
    ⚙️ Процесс обработки
    </div>

    <div style="
        font-size: 0.9rem;
        line-height: 1.6;
        color: var(--text-color);
    ">

    <b>1.</b> Анализ текста — обработка входных обращений<br>
    <b>2.</b> LLM генерация — формирование ответа<br>
    <b>3.</b> Финальный ответ — оформление официального шаблона

    </div>

    </div>
    """, unsafe_allow_html=True)

    ws = get_workspace()
    classified_file_path = ws.get("classified_file")

    if classified_file_path is None:
        st.info("⬅️ Выберите классифицированный файл в правой панели Workspace (папка «Классифицированные»).")
    else:
        st.success(f"✅ Выбран файл: `{os.path.basename(classified_file_path)}`")

        if st.button("▶️ Запустить генерацию", key="generator_run"):
            try:
                st.info("⏳ Загрузка данных...")
                
                df = load_excel_simple(classified_file_path)
                df_test = df.head(5).copy()
                
                st.success("✅ Данные загружены")
                
                st.subheader("📋 Полный тестовый датасет")
                st.dataframe(df_test, use_container_width=True)
            
                df_test["Текст инцидента"] = df_test["Текст инцидента"].apply(clean_text)
                df_test["Группа тем"] = df_test["Группа тем"].fillna("Общее") if "Группа тем" in df_test.columns else "Общее"
                df_test["Тема"] = df_test["Тема"].fillna("Обращение") if "Тема" in df_test.columns else "Обращение"
                df_test["Отдел"] = df_test["Отдел"].fillna("Администрация") if "Отдел" in df_test.columns else "Администрация"
                df_test["Муниципалитет"] = df_test["Муниципалитет"].fillna("Омская область") if "Муниципалитет" in df_test.columns else "Омская область"
                
                st.markdown("""
                <div class="pipeline-title">
                    ⚙️ Процесс генерации
                </div>

                <div class="pipeline">

                    <div class="step-card">
                        <div class="step-number">1</div>
                        <div class="step-content">
                            <b>Подготовка данных</b>
                            <div class="step-desc">Очистка и нормализация обращений</div>
                        </div>
                    </div>

                    <div class="step-card">
                        <div class="step-number">2</div>
                        <div class="step-content">
                            <b>Формирование промпта</b>
                            <div class="step-desc">Добавление темы, отдела и контекста</div>
                        </div>
                    </div>

                    <div class="step-card">
                        <div class="step-number">3</div>
                        <div class="step-content">
                            <b>Генерация LLM</b>
                            <div class="step-desc">Создание официального ответа</div>
                        </div>
                    </div>

                </div>
                """, unsafe_allow_html=True)
                results = []
                progress_bar = st.progress(0)
                
                for idx, (_, row) in enumerate(df_test.iterrows()):
                    progress = (idx + 1) / len(df_test)
                    progress_bar.progress(progress)
                    
                    complaint = str(row["Текст инцидента"])
                    topic = str(row.get("Группа тем", "Общее"))
                    subtopic = str(row.get("Тема", "Обращение"))
                    dept = str(row.get("Отдел", "Администрация"))
                    municipality = str(row.get("Муниципалитет", "Омская область"))
                    
                    real_response = str(row.get("1ый ответ ПИ", ""))
                    if pd.isna(real_response) or real_response == "nan":
                        real_response = ""
                    real_response = clean_text(real_response)
                    
                    try:
                        prompt = build_prompt(complaint, topic, subtopic, dept, municipality)
                        response = generate(prompt=prompt)
                        response = remove_non_russian(response.strip())
                    except Exception as e:
                        response = f"[Ошибка генерации: {e}]"
                        st.warning(f"⚠️ Ошибка на строке {idx + 1}: {e}")
                    
                    results.append({
                        "id": idx,
                        "complaint": complaint[:300],
                        "generated_response": response,
                        "real_response": real_response
                    })
                
                results_df = pd.DataFrame(results)
                
                if "real_response" in results_df.columns:
                    results_df = results_df.drop(columns=["real_response"])
                elif "id" in results_df.columns:
                    results_df = results_df.drop(columns=["id"])

                base_name = os.path.splitext(os.path.basename(classified_file_path))[0]
                generated_filename = f"{base_name}_responses.csv"
                generated_path = fm.get_output_path("generated", generated_filename)
                
                results_df.to_csv(generated_path, index=False, encoding='utf-8-sig')
                
                set_workspace_file("generated", generated_path)
                
                st.success("🎉 Генерация успешно завершена!")

                if os.path.exists(generated_path):
                    with open(generated_path, "rb") as f:
                        st.download_button(
                            label=f"📥 Скачать результаты генерации ({generated_filename})",
                            data=f.read(),
                            file_name=generated_filename,
                            mime="text/csv"
                        )
                    
                    st.markdown("""
                    <div class="pipeline-title">
                        📄 Результаты генерации
                    </div>
                    """, unsafe_allow_html=True)
                    st.dataframe(results_df, use_container_width=True)
                        
            except Exception as e:
                st.error(f"❌ Критическая ошибка: {e}")
                import traceback
                st.text(traceback.format_exc())
