import streamlit as st
import os
import pandas as pd
import time
import sys
import json
from pathlib import Path
from openpyxl import load_workbook

from run_pipeline import extract_and_clean

from src.llm.batcher import create_batches
from src.llm.stage1_is_problem import classify_is_problem
from src.llm.stage2_topic import classify_topic
from src.llm.stage3_department import add_department_to_csv
from src.preprocessing.cleaner import IncidentDataCleaner

from src.workspace import render_workspace_panel, init_workspace, get_workspace, set_workspace_file, FileManager

st.markdown("""
<style>

.page-header{
    margin-bottom:2rem;
}

.page-badge{
    display:inline-block;
    padding:6px 12px;
    border-radius:999px;
    background:rgba(15,98,254,.08);
    color:#0f62fe;
    font-size:.75rem;
    font-weight:700;
    letter-spacing:.08em;
    margin-bottom:.8rem;
}

.title-row{
    display:flex;
    align-items:center;
    gap:14px;
}

.title-bar{
    width:5px;
    height:44px;
    background:#0f62fe;
    border-radius:3px;
    flex-shrink:0;
}

.page-title{
    font-size:2.2rem;
    font-weight:700;
    margin:0;
    line-height:1.1;
}

.page-subtitle{
    margin-top:.8rem;
    color:gray;
    font-size:1rem;
    padding-bottom:1rem;
    border-bottom:1px solid rgba(120,120,120,.2);
}

.block-card{
    border:1px solid rgba(15,98,254,.12);
    border-radius:12px;
    padding:1.25rem;
    background:var(--secondary-background-color);
}

</style>
""", unsafe_allow_html=True)

PROJECT_ROOT = Path(__file__).resolve().parent.parent   
sys.path.insert(0, str(PROJECT_ROOT))

st.set_page_config(page_title="Classificator | Анализатор инцидентов", layout="wide")
st.markdown("""
<div class="page-header">

<div class="page-badge">
CLASSIFICATION MODULE
</div>

<div class="title-row">
    <div class="title-bar"></div>
    <h1 class="page-title">
        Классификатор инцидентов
    </h1>
</div>

<div class="page-subtitle">
Многоэтапная LLM-классификация: определение проблемы, темы и подразделения
</div>

</div>
""", unsafe_allow_html=True)

init_workspace()

left_col, right_col = st.columns([3, 1])

with right_col:
    render_workspace_panel(active_page="classificator")

with left_col:
    fm = FileManager()

    def load_excel_simple(file_path: str) -> pd.DataFrame:
        """Загружает Excel файл с любой структурой столбцов"""
        wb = load_workbook(file_path, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(x).strip() if x else "" for x in rows[0]]
        data = []
        for row in rows[1:]:
            record = {h: row[i] if i < len(row) else None for i, h in enumerate(headers)}
            data.append(record)
        return pd.DataFrame(data)

    def prepare_classifier_data(df: pd.DataFrame) -> pd.DataFrame:
        """
        Преобразует датасет в формат, совместимый с LLM классификатором.
        Если есть только "Текст инцидента", добавляет необходимые столбцы.
        """
        result_df = df.copy()
        
        if "Текст инцидента" not in result_df.columns:
            st.error("❌ Датасет должен содержать столбец 'Текст инцидента'")
            return None
        
        required_cols = ["Отдел", "Тема", "Муниципалитет", "Тип инцидента"]
        for col in required_cols:
            if col not in result_df.columns:
                result_df[col] = ""
        
        if "incident_id" not in result_df.columns:
            result_df["incident_id"] = range(len(result_df))
        
        result_df["Текст инцидента"] = result_df["Текст инцидента"].fillna("").astype(str).str.strip()
        
        return result_df

    def safe_parse_llm_result(result):
        if isinstance(result, str):
            try:
                parsed = json.loads(result)
                if isinstance(parsed, list):
                    return parsed
                elif isinstance(parsed, dict):
                    return [parsed]
            except (json.JSONDecodeError, TypeError):
                pass
        elif isinstance(result, list):
            return result
        return []

    st.markdown("""
    <div style="
        padding: 1.2rem 1.4rem;
        border: 1px solid rgba(15,98,254,.22);
        border-left: 5px solid #0f62fe;
        border-radius: 10px;
        background: var(--secondary-background-color);
        margin: 1rem 0 1.5rem 0;
    ">

    <div style="
        font-size: 1.05rem;
        font-weight: 700;
        margin-bottom: 0.8rem;
    ">
    ⚙️ Пайплайн обработки
    </div>

    <div style="
        font-size: 0.9rem;
        line-height: 1.6;
        color: var(--text-color);
    ">

    <b>1.</b> Этап 1 — определение <b>is_problem</b><br>
    <b>2.</b> Этап 2 — определение <b>topic</b><br>
    <b>3.</b> Этап 3 — определение <b>department</b>

    </div>

    </div>
    """, unsafe_allow_html=True)

    ws = get_workspace()
    raw_file_path = ws.get("raw_file")

    if raw_file_path is None:
        st.info("⬅️ Выберите исходный файл в правой панели Workspace (папка «Исходные файлы»).")
    else:
        st.success(f"✅ Выбран файл: `{os.path.basename(raw_file_path)}`")

        if st.button("▶️ Запустить классификацию", key="classifier_run"):
            try:
                st.info("⏳ Этап 0: Загрузка данных...")
                
                df = load_excel_simple(raw_file_path)
                st.write(f"Загружено строк: {len(df)}")
                
                df = prepare_classifier_data(df)
                if df is None:
                    raise ValueError("Невозможно подготовить данные")
                
                st.success("✅ Данные загружены и подготовлены")
                
                st.subheader("📋 Полный тестовый датасет")
                st.dataframe(df[["Текст инцидента"]], use_container_width=True)

                df_for_records = df.head(20).copy()
                if "incident_id" not in df_for_records.columns:
                    df_for_records["incident_id"] = range(len(df_for_records))
                records = df_for_records.to_dict(orient="records")
                
                st.info("⏳ Этап 1: Определение is_problem...")
                all_problems = []
                progress_bar_1 = st.progress(0)

                batches_list = list(create_batches(records, batch_size=5))
                for batch_number, batch in enumerate(batches_list, start=1):
                    progress = batch_number / len(batches_list)
                    progress_bar_1.progress(progress)

                    try:
                        result = classify_is_problem(batch)
                        parsed = safe_parse_llm_result(result)

                        for i, item in enumerate(parsed):
                            if "incident_id" not in item:
                                item["incident_id"] = batch[i].get("incident_id", i)
                        all_problems.extend(parsed)

                    except Exception as e:
                        st.warning(f"⚠️ Ошибка батча {batch_number}: {e}")
                        for i, row in enumerate(batch):
                            all_problems.append({
                                "incident_id": row.get("incident_id", i),
                                "is_problem": 1
                            })

                if all_problems:
                    problem_df = pd.DataFrame(all_problems)
                else:
                    problem_df = pd.DataFrame(columns=["incident_id", "is_problem"])

                if "incident_id" not in problem_df.columns:
                    problem_df = pd.DataFrame({
                        "incident_id": [rec["incident_id"] for rec in records],
                        "is_problem": 1
                    })

                problem_df = problem_df.drop_duplicates(subset=["incident_id"], keep="first")
                df_merged = pd.DataFrame(records).merge(problem_df, on="incident_id", how="left")
                df_merged["is_problem"] = df_merged["is_problem"].fillna(1).astype(int)
                
                st.info("⏳ Этап 2: Определение темы каждой записи...")
                all_topics = []
                progress_bar_2 = st.progress(0)
                
                batches_list = list(create_batches(records, batch_size=5))
                for batch_number, batch in enumerate(batches_list, start=1):
                    progress = batch_number / len(batches_list)
                    progress_bar_2.progress(progress)
                    
                    try:
                        result = classify_topic(batch)
                        parsed_result = safe_parse_llm_result(result)
                        for i, item in enumerate(parsed_result):
                            if "incident_id" not in item and i < len(batch):
                                item["incident_id"] = batch[i].get("incident_id", i)
                        all_topics.extend(parsed_result)
                        time.sleep(1)
                    except Exception as e:
                        st.warning(f"⚠️ Ошибка батча {batch_number}: {e}")
                        for i, row in enumerate(batch):
                            all_topics.append({
                                "incident_id": row.get("incident_id", i),
                                "topic": "Прочее"
                            })
                
                topic_df = pd.DataFrame(all_topics)
                topic_df = topic_df.drop_duplicates(subset=["incident_id"], keep="first")
                
                df_merged = df_merged.merge(topic_df, on="incident_id", how="left")
                df_merged['topic'] = df_merged['topic'].fillna("Прочее")
                
                st.info("⏳ Этап 3: Определение отдела")

                base_name = os.path.splitext(os.path.basename(raw_file_path))[0]
                classified_filename = f"{base_name}_classified.xlsx"
                classified_path = fm.get_output_path("classified", classified_filename)

                temp_topics = fm.get_output_path("classified", "temp_topics.csv")
                temp_csv = fm.get_output_path("classified", "temp_final.csv")
                
                df_merged.to_csv(temp_topics, index=False, encoding="utf-8-sig")
                add_department_to_csv(temp_topics, temp_csv)
                
                result_df = pd.read_csv(temp_csv)
                with pd.ExcelWriter(classified_path, engine="openpyxl") as writer:
                    result_df.to_excel(writer, sheet_name="Результаты", index=False)
                
                set_workspace_file("classified", classified_path)
                
                st.success("🎉 Классификация успешно завершена!")
                
                if os.path.exists(classified_path):
                    result_df = pd.read_excel(classified_path)
                    
                    output_columns = ["Текст инцидента", "is_problem", "topic", "department"]
                    available_columns = [col for col in output_columns if col in result_df.columns]
                    result_df_output = result_df[available_columns]
                    
                    with open(classified_path, "rb") as f:
                        st.download_button(
                            label=f"📥 Скачать результаты классификации ({classified_filename})",
                            data=f.read(),
                            file_name=classified_filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    
                    st.subheader("📋 Превью результатов")
                    st.dataframe(result_df_output, use_container_width=True)
                        
            except Exception as e:
                st.error(f"❌ Критическая ошибка: {e}")
                import traceback
                st.text(traceback.format_exc())
