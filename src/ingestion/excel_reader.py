from openpyxl import load_workbook

from src.ingestion.schema import REQUIRED_COLUMNS


def read_excel(file_path: str):
    wb = load_workbook(file_path, read_only=True)

    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))

    headers = [str(x).strip() if x else "" for x in rows[0]]

    column_indexes = {}

    for col in REQUIRED_COLUMNS:
        if col not in headers:
            raise ValueError(f"Не найден столбец: {col}")

        column_indexes[col] = headers.index(col)

    result = []

    for row in rows[1:]:

        record = {}

        for col_name, idx in column_indexes.items():

            value = row[idx] if idx < len(row) else None

            record[col_name] = value

        result.append(record)

    return result
